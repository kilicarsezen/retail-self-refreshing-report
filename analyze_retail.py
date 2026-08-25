from __future__ import annotations

import argparse
import csv
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


EXPECTED_COLUMNS = (
    "Invoice",
    "StockCode",
    "Description",
    "Quantity",
    "InvoiceDate",
    "Price",
    "Customer ID",
    "Country",
)


def analyze_workbook(input_path: Path, output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    total_rows = 0
    valid_rows = 0
    flagged_rows = 0
    cancelled_rows = 0
    revenue = 0.0
    invoices: set[str] = set()
    customers: set[str] = set()
    monthly_sales: defaultdict[str, float] = defaultdict(float)
    product_sales: Counter[str] = Counter()
    country_sales: Counter[str] = Counter()

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    invalid_file = output_dir / "invalid_rows.csv"
    invalid_output = invalid_file.open("w", newline="", encoding="utf-8")
    invalid_writer = csv.writer(invalid_output)
    invalid_writer.writerow(("sheet", "excel_row", "flag", *EXPECTED_COLUMNS))
    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            columns = tuple(next(rows))
            if columns != EXPECTED_COLUMNS:
                raise ValueError(
                    f"Unexpected columns in {worksheet.title!r}: {columns}"
                )

            for excel_row, row in enumerate(rows, start=2):
                total_rows += 1
                values = list(row) + [None] * (len(EXPECTED_COLUMNS) - len(row))
                invoice, stock_code, description, quantity, invoice_date, price, customer_id, country = values[:8]
                flags = []
                if not invoice:
                    flags.append("missing_invoice")
                if not stock_code:
                    flags.append("missing_stock_code")
                if not invoice_date:
                    flags.append("missing_invoice_date")
                if quantity is None:
                    flags.append("missing_quantity")
                if price is None:
                    flags.append("missing_price")
                if customer_id is None:
                    flags.append("missing_customer_id")

                try:
                    quantity_value = float(quantity) if quantity is not None else None
                    price_value = float(price) if price is not None else None
                except (TypeError, ValueError):
                    quantity_value = price_value = None
                    flags.append("non_numeric_quantity_or_price")

                if flags:
                    flagged_rows += 1
                    invalid_writer.writerow(
                        (worksheet.title, excel_row, ";".join(flags), *values[:8])
                    )

                if not invoice or not stock_code or not invoice_date:
                    continue
                if quantity_value is None or price_value is None:
                    continue

                valid_rows += 1
                invoice_text = str(invoice)
                row_revenue = quantity_value * price_value
                is_cancelled = invoice_text.upper().startswith("C")

                if is_cancelled:
                    cancelled_rows += 1
                else:
                    revenue += row_revenue
                    invoices.add(invoice_text)
                    if customer_id is not None:
                        customers.add(str(customer_id))
                    month = invoice_date.strftime("%Y-%m") if isinstance(invoice_date, datetime) else str(invoice_date)[:7]
                    monthly_sales[month] += row_revenue
                    product_sales[str(description or stock_code)] += row_revenue
                    country_sales[str(country or "Unknown")] += row_revenue
    finally:
        invalid_output.close()
        workbook.close()

    write_csv(output_dir / "monthly_sales.csv", monthly_sales, "month")
    write_csv(output_dir / "top_products.csv", product_sales.most_common(20), "product")
    write_csv(output_dir / "top_countries.csv", country_sales.most_common(20), "country")

    summary = {
        "rows_read": total_rows,
        "valid_rows": valid_rows,
        "revenue": round(revenue, 2),
        "orders": len(invoices),
        "customers": len(customers),
        "cancelled_rows": cancelled_rows,
        "flagged_rows": flagged_rows,
    }
    with (output_dir / "summary.csv").open("w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(("metric", "value"))
        writer.writerows(summary.items())

    print(f"Rows read: {total_rows:,}")
    print(f"Revenue: ${revenue:,.2f}")
    print(f"Orders: {len(invoices):,}")
    print(f"Customers: {len(customers):,}")
    print(f"Cancelled rows: {cancelled_rows:,}")
    print(f"Reports written to: {output_dir}")


def write_csv(path: Path, values, first_column_name: str) -> None:
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow((first_column_name, "revenue"))
        writer.writerows(values)


def main() -> None:
    parser = argparse.ArgumentParser(description="Analyze an Online Retail II workbook.")
    parser.add_argument(
        "input_path",
        nargs="?",
        type=Path,
        default=Path("data/online_retail_II.xlsx"),
        help="Path to the Excel workbook.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("reports"),
        help="Directory for generated CSV reports.",
    )
    args = parser.parse_args()
    analyze_workbook(args.input_path, args.output_dir)


if __name__ == "__main__":
    main()